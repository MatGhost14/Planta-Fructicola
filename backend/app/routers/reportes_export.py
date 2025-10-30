import hashlib
from fastapi.responses import HTMLResponse
def calcular_hash_reporte(inspecciones, filtros):
    """
    Calcula un hash determinista del reporte, usando los datos principales y los hashes de evidencias (fotos).
    """
    manifest = []
    for insp in inspecciones:
        fotos_hashes = sorted([f.hash_hex or '' for f in getattr(insp, 'fotos', [])])
        manifest.append(f"{insp.id_inspeccion}|{insp.codigo}|{insp.numero_contenedor}|{insp.estado}|{insp.inspeccionado_en}|{','.join(fotos_hashes)}")
    manifest_str = '\n'.join(sorted(manifest))
    filtros_str = str(sorted(filtros.items()))
    hash_input = (manifest_str + filtros_str).encode('utf-8')
    return hashlib.sha256(hash_input).hexdigest()

"""
Router para exportación de reportes en PDF y Excel
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date
from typing import Optional
import io

# Importaciones para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
import qrcode
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Importaciones para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..core.database import get_db
from ..models import Inspeccion, Planta, Naviera, Usuario
from ..utils.auth import get_current_active_user

router = APIRouter(tags=["Reportes Export"])


def crear_pdf_inspecciones(inspecciones: list, filtros: dict) -> io.BytesIO:
    """
    Genera un PDF profesional y moderno con el reporte de inspecciones
    """
    try:
        buffer = io.BytesIO()
        # Definir colores corporativos
        COLOR_PRIMARY = colors.HexColor('#2563eb')  # Azul moderno
        COLOR_SUCCESS = colors.HexColor('#10b981')  # Verde (Aprobado)
        COLOR_WARNING = colors.HexColor('#f59e0b')  # Amarillo (Pendiente)
        COLOR_DANGER = colors.HexColor('#ef4444')   # Rojo (Rechazado)
        COLOR_DARK = colors.HexColor('#1e293b')
        COLOR_LIGHT_BG = colors.HexColor('#f8fafc')
        COLOR_BORDER = colors.HexColor('#e2e8f0')
        # Estilos personalizados
        styles = getSampleStyleSheet()
        # Título principal
        titulo_principal = ParagraphStyle(
            'TituloPrincipal',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=COLOR_PRIMARY,
            fontName='Helvetica-Bold',
            spaceAfter=10,
            alignment=TA_CENTER,
            leading=28)
        # Subtítulo
        subtitulo_style = ParagraphStyle(
            'Subtitulo',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#64748b'),
            alignment=TA_CENTER,
            spaceAfter=25,
            leading=14)
        # Título de sección
        titulo_seccion = ParagraphStyle(
            'TituloSeccion',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=COLOR_DARK,
            fontName='Helvetica-Bold',
            spaceAfter=12,
            spaceBefore=20,
            leading=17)

        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        elementos = []

        # ========== ENCABEZADO ========== 
        # Logo/Título de la empresa
        elementos.append(Paragraph("🏭 INSPECCIÓN DE CONTENEDORES", titulo_principal))
        elementos.append(Paragraph("Sistema de Control de Calidad", subtitulo_style))
        # Línea decorativa
        elementos.append(Spacer(1, 0.1*inch))
        line_data = [['', '', '', '']]
        line_table = Table(line_data, colWidths=[2*inch, 2*inch, 2*inch, 1.2*inch])
        line_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, 0), 3, COLOR_PRIMARY),
        ]))
        elementos.append(line_table)
        elementos.append(Spacer(1, 0.2*inch))

        # ========== CÓDIGO QR DE VERIFICACIÓN ========== 
        id_reporte = inspecciones[0].id_inspeccion if inspecciones else 0
        try:
            hash_reporte = calcular_hash_reporte(inspecciones, filtros)
        except Exception:
            hash_reporte = "errorhash"
        url_verificacion = f"https://tuservidor/verificar-reporte/{id_reporte}/{hash_reporte}"
        try:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=4,
                border=2,
            )
            qr.add_data(url_verificacion)
            qr.make(fit=True)
            img_qr = qr.make_image(fill_color="black", back_color="white")
            qr_buffer = io.BytesIO()
            img_qr.save(qr_buffer, format="PNG")
            qr_buffer.seek(0)
            qr_img = Image(qr_buffer, width=1*inch, height=1*inch)
            qr_table = Table([[qr_img]], colWidths=[1*inch], rowHeights=[1*inch])
            qr_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elementos.append(qr_table)
            elementos.append(Spacer(1, 0.1*inch))
        except Exception:
            pass

        # ========== INFORMACIÓN DEL REPORTE ========== 
        fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
        info_data = [
            ['📅 Fecha de Generación:', fecha_actual, '📊 Total de Registros:', str(len(inspecciones))]
        ]
        if filtros.get('fecha_desde') and filtros.get('fecha_hasta'):
            info_data.append([
                '📆 Período:', 
                f"{filtros['fecha_desde']} al {filtros['fecha_hasta']}", 
                '', 
                ''
            ])
        info_table = Table(info_data, colWidths=[1.8*inch, 2.5*inch, 1.8*inch, 1.1*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), COLOR_LIGHT_BG),
            ('TEXTCOLOR', (0, 0), (-1, -1), COLOR_DARK),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('BOX', (0, 0), (-1, -1), 1, COLOR_BORDER),
        ]))
        elementos.append(info_table)
        elementos.append(Spacer(1, 0.3*inch))

        # ========== RESUMEN ESTADÍSTICO ========== 
        elementos.append(Paragraph("📈 RESUMEN ESTADÍSTICO", titulo_seccion))
        elementos.append(Spacer(1, 0.15*inch))
        total = len(inspecciones)
        aprobadas = sum(1 for i in inspecciones if i.estado == 'approved')
        rechazadas = sum(1 for i in inspecciones if i.estado == 'rejected')
        pendientes = sum(1 for i in inspecciones if i.estado == 'pending')
        porcentaje_aprobadas = (aprobadas / total * 100) if total > 0 else 0
        porcentaje_rechazadas = (rechazadas / total * 100) if total > 0 else 0
        porcentaje_pendientes = (pendientes / total * 100) if total > 0 else 0
        resumen_data = [
            ['ESTADO', 'CANTIDAD', 'PORCENTAJE'],
            ['OK Aprobadas', str(aprobadas), f'{porcentaje_aprobadas:.1f}%'],
            ['Pendientes', str(pendientes), f'{porcentaje_pendientes:.1f}%'],
            ['X Rechazadas', str(rechazadas), f'{porcentaje_rechazadas:.1f}%'],
            ['TOTAL', str(total), '100%']
        ]
        resumen_table = Table(resumen_data, colWidths=[3*inch, 1.8*inch, 1.8*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecfdf5')),
            ('TEXTCOLOR', (0, 1), (-1, 1), COLOR_SUCCESS),
            ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#fffbeb')),
            ('TEXTCOLOR', (0, 2), (-1, 2), COLOR_WARNING),
            ('FONTNAME', (0, 2), (0, 2), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fef2f2')),
            ('TEXTCOLOR', (0, 3), (-1, 3), COLOR_DANGER),
            ('FONTNAME', (0, 3), (0, 3), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 4), (-1, 4), COLOR_DARK),
            ('TEXTCOLOR', (0, 4), (-1, 4), colors.white),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
        ]))
        elementos.append(resumen_table)
        elementos.append(Spacer(1, 0.4*inch))

        # ========== DETALLE DE INSPECCIONES ========== 
        if inspecciones:
            elementos.append(Paragraph("DETALLE DE INSPECCIONES", titulo_seccion))
            elementos.append(Spacer(1, 0.15*inch))
            data = [['CÓDIGO', 'CONTENEDOR', 'PLANTA', 'FECHA', 'ESTADO', 'INSPECTOR']]
            for insp in inspecciones[:100]:
                estado_texto = {
                    'pending': '⏳ Pendiente',
                    'approved': 'OK Aprobado',
                    'rejected': 'X Rechazado'
                }.get(insp.estado, insp.estado)
                fecha = insp.inspeccionado_en.strftime("%d/%m/%Y") if insp.inspeccionado_en else 'N/A'
                planta_nombre = insp.planta.nombre if insp.planta else 'N/A'
                inspector_nombre = insp.inspector.nombre if insp.inspector else 'N/A'
                data.append([
                    insp.codigo[:18] if insp.codigo else 'N/A',
                    insp.numero_contenedor[:15] if insp.numero_contenedor else 'N/A',
                    planta_nombre[:18],
                    fecha,
                    estado_texto,
                    inspector_nombre[:18]
                ])
            col_widths = [1.3*inch, 1.4*inch, 1.3*inch, 0.9*inch, 1.1*inch, 1.2*inch]
            tabla = Table(data, colWidths=col_widths)
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (5, -1), 'LEFT'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_BG]),
            ]
            for i, insp in enumerate(inspecciones[:100], start=1):
                if insp.estado == 'approved':
                    table_style.append(('TEXTCOLOR', (4, i), (4, i), COLOR_SUCCESS))
                    table_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
                elif insp.estado == 'rejected':
                    table_style.append(('TEXTCOLOR', (4, i), (4, i), COLOR_DANGER))
                    table_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
                elif insp.estado == 'pending':
                    table_style.append(('TEXTCOLOR', (4, i), (4, i), COLOR_WARNING))
                    table_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
            tabla.setStyle(TableStyle(table_style))
            elementos.append(tabla)
            elementos.append(Spacer(1, 0.3*inch))
            pie_style = ParagraphStyle(
                'Pie',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#94a3b8'),
                alignment=TA_CENTER
            )
            elementos.append(Paragraph(
                "Este reporte ha sido generado automáticamente por el Sistema de Inspección de Contenedores",
                pie_style
            ))
            elementos.append(Paragraph(
                f"Hash de verificación: {hash_reporte[:16]}...",
                pie_style
            ))

        doc.build(elementos)
        buffer.seek(0)
        return buffer
    except Exception as e:
        buffer_err = io.BytesIO()
        doc = SimpleDocTemplate(buffer_err, pagesize=A4)
        styles = getSampleStyleSheet()
        doc.build([Paragraph(f"Error generando PDF: {str(e)}", styles['Normal'])])
        buffer_err.seek(0)
        return buffer_err
    info_data = [
        ['📅 Fecha de Generación:', fecha_actual, '📊 Total de Registros:', str(len(inspecciones))]
    ]
    
    if filtros.get('fecha_desde') and filtros.get('fecha_hasta'):
        info_data.append([
            '📆 Período:', 
            f"{filtros['fecha_desde']} al {filtros['fecha_hasta']}", 
            '', 
            ''
        ])
    
    info_table = Table(info_data, colWidths=[1.8*inch, 2.5*inch, 1.8*inch, 1.1*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COLOR_LIGHT_BG),
        ('TEXTCOLOR', (0, 0), (-1, -1), COLOR_DARK),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('BOX', (0, 0), (-1, -1), 1, COLOR_BORDER),
    ]))
    elementos.append(info_table)
    elementos.append(Spacer(1, 0.3*inch))
    
    # ========== RESUMEN ESTADÍSTICO ==========
    elementos.append(Paragraph("📈 RESUMEN ESTADÍSTICO", titulo_seccion))
    elementos.append(Spacer(1, 0.15*inch))
    
    total = len(inspecciones)
    aprobadas = sum(1 for i in inspecciones if i.estado == 'approved')
    rechazadas = sum(1 for i in inspecciones if i.estado == 'rejected')
    pendientes = sum(1 for i in inspecciones if i.estado == 'pending')
    
    porcentaje_aprobadas = (aprobadas / total * 100) if total > 0 else 0
    porcentaje_rechazadas = (rechazadas / total * 100) if total > 0 else 0
    porcentaje_pendientes = (pendientes / total * 100) if total > 0 else 0
    
    resumen_data = [
        ['ESTADO', 'CANTIDAD', 'PORCENTAJE'],
        ['OK Aprobadas', str(aprobadas), f'{porcentaje_aprobadas:.1f}%'],
        ['Pendientes', str(pendientes), f'{porcentaje_pendientes:.1f}%'],
        ['X Rechazadas', str(rechazadas), f'{porcentaje_rechazadas:.1f}%'],
        ['TOTAL', str(total), '100%']
    ]
    
    resumen_table = Table(resumen_data, colWidths=[3*inch, 1.8*inch, 1.8*inch])
    resumen_table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Fila Aprobadas
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ecfdf5')),
        ('TEXTCOLOR', (0, 1), (-1, 1), COLOR_SUCCESS),
        ('FONTNAME', (0, 1), (0, 1), 'Helvetica-Bold'),
        
        # Fila Pendientes
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#fffbeb')),
        ('TEXTCOLOR', (0, 2), (-1, 2), COLOR_WARNING),
        ('FONTNAME', (0, 2), (0, 2), 'Helvetica-Bold'),
        
        # Fila Rechazadas
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fef2f2')),
        ('TEXTCOLOR', (0, 3), (-1, 3), COLOR_DANGER),
        ('FONTNAME', (0, 3), (0, 3), 'Helvetica-Bold'),
        
        # Fila Total
        ('BACKGROUND', (0, 4), (-1, 4), COLOR_DARK),
        ('TEXTCOLOR', (0, 4), (-1, 4), colors.white),
        ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
        
        # General
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
    ]))
    
    elementos.append(resumen_table)
    elementos.append(Spacer(1, 0.4*inch))
    
    # ========== DETALLE DE INSPECCIONES ==========
    if inspecciones:
        elementos.append(Paragraph("DETALLE DE INSPECCIONES", titulo_seccion))
        elementos.append(Spacer(1, 0.15*inch))
        
        # Encabezados
        data = [['CÓDIGO', 'CONTENEDOR', 'PLANTA', 'FECHA', 'ESTADO', 'INSPECTOR']]
        
        # Datos con formato condicional
        for insp in inspecciones[:100]:  # Limitar a 100 registros
            estado_texto = {
                'pending': '⏳ Pendiente',
                'approved': 'OK Aprobado',
                'rejected': 'X Rechazado'
            }.get(insp.estado, insp.estado)
            
            fecha = insp.inspeccionado_en.strftime("%d/%m/%Y") if insp.inspeccionado_en else 'N/A'
            planta_nombre = insp.planta.nombre if insp.planta else 'N/A'
            inspector_nombre = insp.inspector.nombre if insp.inspector else 'N/A'
            
            data.append([
                insp.codigo[:18] if insp.codigo else 'N/A',
                insp.numero_contenedor[:15] if insp.numero_contenedor else 'N/A',
                planta_nombre[:18],
                fecha,
                estado_texto,
                inspector_nombre[:18]
            ])
        
        # Crear tabla con mejor diseño
        col_widths = [1.3*inch, 1.4*inch, 1.3*inch, 0.9*inch, 1.1*inch, 1.2*inch]
        tabla = Table(data, colWidths=col_widths)
        
        # Estilo moderno para la tabla
        table_style = [
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Contenido
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (5, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            
            # Bordes y fondos alternados
            ('GRID', (0, 0), (-1, -1), 0.5, COLOR_BORDER),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_BG]),
        ]
        
        # Aplicar colores según estado
        for i, insp in enumerate(inspecciones[:100], start=1):
            if insp.estado == 'approved':
                table_style.append(('TEXTCOLOR', (4, i), (4, i), COLOR_SUCCESS))
                table_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
            elif insp.estado == 'rejected':
                table_style.append(('TEXTCOLOR', (4, i), (4, i), COLOR_DANGER))
                table_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
            elif insp.estado == 'pending':
                table_style.append(('TEXTCOLOR', (4, i), (4, i), COLOR_WARNING))
                table_style.append(('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'))
        
        tabla.setStyle(TableStyle(table_style))
        elementos.append(tabla)
        
        # Pie de página
        elementos.append(Spacer(1, 0.3*inch))
        pie_style = ParagraphStyle(
            'Pie',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        elementos.append(Paragraph(
            "Este reporte ha sido generado automáticamente por el Sistema de Inspección de Contenedores",
            pie_style
        ))
        elementos.append(Paragraph(
            f"Hash de verificación: {hash_reporte[:16]}...",
            pie_style
        ))
@router.get("/verificar-reporte/{id_reporte}/{hash_reporte}", response_class=HTMLResponse)
def verificar_reporte(id_reporte: int, hash_reporte: str, db: Session = Depends(get_db)):
    # Buscar todas las inspecciones con el mismo id_reporte (puede ser un lote, aquí solo una)
    inspecciones = db.query(Inspeccion).filter(Inspeccion.id_inspeccion == id_reporte).all()
    if not inspecciones:
        return HTMLResponse("<h2>Reporte no encontrado</h2>", status_code=404)
    # Para demo, usar filtros vacíos (en producción, reconstruir los filtros usados)
    filtros = {}
    hash_actual = calcular_hash_reporte(inspecciones, filtros)
    if hash_actual == hash_reporte:
        estado = "<span style='color:green;font-weight:bold'>VÁLIDO</span>"
    else:
        estado = "<span style='color:red;font-weight:bold'>ALTERADO</span>"
    html = f"""
    <html><head><title>Verificación de Reporte</title></head><body>
    <h2>Verificación de Reporte #{id_reporte}</h2>
    <p>Estado: {estado}</p>
    <p>Hash esperado: <code>{hash_reporte}</code></p>
    <p>Hash actual: <code>{hash_actual}</code></p>
    </body></html>
    """
    return HTMLResponse(html)


def crear_excel_inspecciones(inspecciones: list, filtros: dict) -> io.BytesIO:
    """
    Genera un archivo Excel con el reporte de inspecciones
    """
    buffer = io.BytesIO()
    wb = Workbook()
    
    # Hoja de Resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Título
    ws_resumen['A1'] = 'REPORTE DE INSPECCIONES DE CONTENEDORES'
    ws_resumen['A1'].font = Font(size=16, bold=True, color='1e3a8a')
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A1:F1')
    
    # Información del reporte
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws_resumen['A2'] = f'Generado el: {fecha_actual}'
    ws_resumen['A2'].alignment = Alignment(horizontal='center')
    ws_resumen.merge_cells('A2:F2')
    
    if filtros.get('fecha_desde') and filtros.get('fecha_hasta'):
        ws_resumen['A3'] = f"Período: {filtros['fecha_desde']} al {filtros['fecha_hasta']}"
        ws_resumen['A3'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A3:F3')
    
    # Estadísticas
    row = 5
    ws_resumen[f'A{row}'] = 'ESTADÍSTICAS GENERALES'
    ws_resumen[f'A{row}'].font = Font(size=12, bold=True)
    ws_resumen.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ['Total Inspecciones', 'Aprobadas', 'Rechazadas', 'Pendientes']
    for col, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='e5e7eb', end_color='e5e7eb', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    row += 1
    total = len(inspecciones)
    aprobadas = sum(1 for i in inspecciones if i.estado == 'approved')
    rechazadas = sum(1 for i in inspecciones if i.estado == 'rejected')
    pendientes = sum(1 for i in inspecciones if i.estado == 'pending')
    
    valores = [total, aprobadas, rechazadas, pendientes]
    for col, valor in enumerate(valores, 1):
        cell = ws_resumen.cell(row, col, valor)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Ajustar anchos de columna
    for col in range(1, 7):
        ws_resumen.column_dimensions[get_column_letter(col)].width = 20
    
    # Hoja de Detalle
    ws_detalle = wb.create_sheet(title="Detalle")
    
    # Encabezados
    encabezados = ['Código', 'N° Contenedor', 'Planta', 'Naviera', 'Fecha Inspección', 'Estado', 'Inspector', 'Observaciones']
    for col, header in enumerate(encabezados, 1):
        cell = ws_detalle.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Datos
    for row_idx, insp in enumerate(inspecciones, 2):
        estado_texto = {
            'pending': 'Pendiente',
            'approved': 'Aprobado',
            'rejected': 'Rechazado'
        }.get(insp.estado, insp.estado)
        
        fecha = insp.inspeccionado_en.strftime("%d/%m/%Y") if insp.inspeccionado_en else 'N/A'
        planta_nombre = insp.planta.nombre if insp.planta else 'N/A'
        naviera_nombre = insp.naviera.nombre if insp.naviera else 'N/A'
        inspector_nombre = insp.inspector.nombre if insp.inspector else 'N/A'
        
        datos = [
            insp.codigo or 'N/A',
            insp.numero_contenedor or 'N/A',
            planta_nombre,
            naviera_nombre,
            fecha,
            estado_texto,
            inspector_nombre,
            insp.observaciones[:50] if insp.observaciones else ''
        ]
        
        for col_idx, valor in enumerate(datos, 1):
            cell = ws_detalle.cell(row_idx, col_idx, valor)
            cell.alignment = Alignment(horizontal='left' if col_idx in [1, 2, 3, 4, 7, 8] else 'center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color según estado
            if col_idx == 6:  # Columna Estado
                if estado_texto == 'Aprobado':
                    cell.fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
                    cell.font = Font(color='065f46', bold=True)
                elif estado_texto == 'Rechazado':
                    cell.fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
                    cell.font = Font(color='991b1b', bold=True)
                elif estado_texto == 'Pendiente':
                    cell.fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
                    cell.font = Font(color='92400e', bold=True)
    
    # Ajustar anchos de columna
    anchos = [15, 18, 25, 25, 18, 15, 25, 40]
    for col, ancho in enumerate(anchos, 1):
        ws_detalle.column_dimensions[get_column_letter(col)].width = ancho
    
    # Guardar
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/pdf")
async def exportar_pdf(
    fecha_desde: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    estado: Optional[str] = Query(None, description="Estado: pending, approved, rejected"),
    id_planta: Optional[int] = Query(None, description="ID de la planta"),
    id_inspector: Optional[int] = Query(None, description="ID del inspector"),
    current_user: Usuario = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Genera un reporte en PDF con las inspecciones filtradas
    """
    try:
        # Construir query base (sin JOINs, usaremos las relaciones de SQLAlchemy)
        query = db.query(Inspeccion)
        
        # Filtrar por rol
        if current_user.rol == 'inspector':
            query = query.filter(Inspeccion.id_inspector == current_user.id_usuario)
        
        # Filtros adicionales
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
                query = query.filter(func.date(Inspeccion.inspeccionado_en) >= fecha_desde_obj)
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido")
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
                query = query.filter(func.date(Inspeccion.inspeccionado_en) <= fecha_hasta_obj)
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido")
        
        if estado:
            query = query.filter(Inspeccion.estado == estado)
        
        if id_planta:
            query = query.filter(Inspeccion.id_planta == id_planta)
        
        if id_inspector:
            query = query.filter(Inspeccion.id_inspector == id_inspector)
        
        # Ordenar y obtener resultados
        inspecciones = query.order_by(Inspeccion.inspeccionado_en.desc()).limit(1000).all()
        
        if not inspecciones:
            raise HTTPException(status_code=404, detail="No se encontraron inspecciones con los filtros aplicados")
        
        # Generar PDF
        filtros = {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'estado': estado,
            'id_planta': id_planta,
            'id_inspector': id_inspector
        }
        
        pdf_buffer = crear_pdf_inspecciones(inspecciones, filtros)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_inspecciones_{timestamp}.pdf"
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")


@router.get("/excel")
async def exportar_excel(
    fecha_desde: Optional[str] = Query(None, description="Fecha desde (YYYY-MM-DD)"),
    fecha_hasta: Optional[str] = Query(None, description="Fecha hasta (YYYY-MM-DD)"),
    estado: Optional[str] = Query(None, description="Estado: pending, approved, rejected"),
    id_planta: Optional[int] = Query(None, description="ID de la planta"),
    id_inspector: Optional[int] = Query(None, description="ID del inspector"),
    current_user: Usuario = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Genera un reporte en Excel con las inspecciones filtradas
    """
    try:
        # Construir query base (sin JOINs, usaremos las relaciones de SQLAlchemy)
        query = db.query(Inspeccion)
        
        # Filtrar por rol
        if current_user.rol == 'inspector':
            query = query.filter(Inspeccion.id_inspector == current_user.id_usuario)
        
        # Filtros adicionales
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
                query = query.filter(func.date(Inspeccion.inspeccionado_en) >= fecha_desde_obj)
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_desde inválido")
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
                query = query.filter(func.date(Inspeccion.inspeccionado_en) <= fecha_hasta_obj)
            except ValueError:
                raise HTTPException(status_code=400, detail="Formato de fecha_hasta inválido")
        
        if estado:
            query = query.filter(Inspeccion.estado == estado)
        
        if id_planta:
            query = query.filter(Inspeccion.id_planta == id_planta)
        
        if id_inspector:
            query = query.filter(Inspeccion.id_inspector == id_inspector)
        
        # Ordenar y obtener resultados
        inspecciones = query.order_by(Inspeccion.inspeccionado_en.desc()).limit(5000).all()
        
        if not inspecciones:
            raise HTTPException(status_code=404, detail="No se encontraron inspecciones con los filtros aplicados")
        
        # Generar Excel
        filtros = {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'estado': estado,
            'id_planta': id_planta,
            'id_inspector': id_inspector
        }
        
        excel_buffer = crear_excel_inspecciones(inspecciones, filtros)
        
        # Nombre del archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_inspecciones_{timestamp}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")
